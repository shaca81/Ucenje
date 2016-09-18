import openpyxl
from qrcode import *

wb = openpyxl.load_workbook('BOM.xlsx')
sheet = wb.get_sheet_by_name('PWT TOP - Bill of Material')

qr = QRCode(
    version=3,
    error_correction=ERROR_CORRECT_L,
    box_size=10,
    border=2)

for i in range(4, 63):
    print(str(sheet.cell(row=i, column=1).value) + "\n " + str(sheet.cell(row=i, column=3).value) + "\n " \
                                                                                                    "" + str(
        sheet.cell(row=i, column=4).value) + "\n\n")

    qr.add_data(str(sheet.cell(row=i, column=1).value) + "\n " + str(sheet.cell(row=i, column=3).value) + "\n " \
                                                                                                          "" + str(
        sheet.cell(row=i, column=4).value) + "\n\n")

    qr.make()  # Generate the QRCode itself
    #  im contains a PIL.Image.Image object
    image = qr.make_image()  # To save it
    image.save(str(sheet.cell(row=i, column=1).value) + '.png')
    qr.clear()
